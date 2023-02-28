import openpyxl


# 读取全局变量工作表
def get_variable(wb):
    sheet_data = wb["全局变量"]
    variable = {}  # 用来储存读到的变量，名称是key，值是value
    lines_count = sheet_data.max_row  # 获取总行数
    for l in range(2, lines_count+1):
        key = sheet_data.cell(l, 1).value
        value = sheet_data.cell(l, 2).value
        variable[key] = value
    return variable

# 读取接口默认参数
def get_api_default_params(wb):
    sheet_data = wb["接口默认参数"]
    api_default_params = {}  # 用来储存读到的变量，名称是key，值是value
    lines_count = sheet_data.max_row  # 获取总行数
    for l in range(2, lines_count + 1):
        key = sheet_data.cell(l, 1).value
        value = sheet_data.cell(l, 2).value
        api_default_params[key] = value
    return api_default_params


# 获取要执行的测试集合名称
def get_casesuitename(wb):
    sheet_data = wb["测试集合管理"]
    case_suite_name = []  # 用来储存读到的变量，名称是key，值是value
    lines_count = sheet_data.max_row  # 获取总行数
    for l in range(2, lines_count + 1):
        flag = sheet_data.cell(l, 2).value
        if flag == "y":
            key = sheet_data.cell(l, 1).value
            case_suite_name.append(key)
    return case_suite_name

# 需要根据要执行的测试集合名称来读取对应的测试用例数据
def read_testcases(wb,suite_name):
    sheet_data = wb[suite_name]
    lines_count = sheet_data.max_row  # 获取总行数
    cols_count = sheet_data.max_column  # 获取总列数
    """
    规定读出来的测试数据存储结构如下：
    {
    “新增客户正确”:[
    ['apiname','接口地址','请求方式','头信息',	],
    ['apiname','接口地址','请求方式','头信息',	],
    ],
    " 新 增 客 户 失 败 - 用 户 名 为 空 ":[ ['apiname','接口地址','请求方式','头信息',	]
    ],
    "新增客户失败-手机号格式不正确":[
    ['apiname','接口地址','请求方式','头信息',	]
    ]
    } """

    cases_info = {}  # 用来存储当前测试集合中的所有用例信息的
    for l in range(2, lines_count + 1):
        case_name = sheet_data.cell(l, 2).value  # 测试用例名称
        lines = []  # 用来存储当前行测试数据的
        for c in range(3, cols_count + 1):
            cell = sheet_data.cell(l, c).value  # 当前单元格数据
            if cell == None:  # 处理空单元格
                cell = ''
            lines.append(cell)
    # 判断当前用例名称是否已存在于cases_info中
    # 如果不存在，那就是直接赋值
    # 否则就是在原来的基础上追加
        if case_name not in cases_info:
            cases_info[case_name] = [lines]
        else:
            cases_info[case_name].append(lines)
    return cases_info


# 整合所有要执行的测试用例数据，将其转成pytest参数化需要的数据结构格式
def get_all_testcases(wb):
    """
    整合后的数据结构是
    [
    ['新增客户接口测试集合','新增客户正确',[[],[]]],
    ['新增客户接口测试集合','新增客户失败-用户名为空',[[],[]]],
    ['新增客户接口测试集合','新增客户失败-手机号格式不正确',[[],[]]],
    ['新建产品接口测试集合','新建产品正确',[[],[]]],
    ['新建产品接口测试集合','新建产品失败-产品编码重复',[[],[]]],
    ]
    # :param wb:
    # :return:
    """
    test_data = [] # 用来存储所有测试数据
    # 获取所有要执行的测试集合名称
    cases_suite_name = get_casesuitename(wb)
    for suite_name in cases_suite_name:
        # 遍历读取每个要执行的测试集合sheet工作表中的测试用例数据
        cur_cases_info = read_testcases(wb, suite_name) # 是个字典
        for key, value in cur_cases_info.items():
            # key实际上就是测试用例名称，value实际上测试用例多行数据信息
            case_info = [suite_name,key,value]
            test_data.append(case_info)
    return test_data


if __name__ == '__main__':
    wb = openpyxl.load_workbook('../testcases/CRM系统接口测试用例.xlsx')
    # print(get_variable(wb))
    # print(get_api_default_params(wb))
    # print(get_casesuitename(wb))
    # print(read_testcases(wb,'新增客户接口测试集合'))
    print(get_all_testcases(wb))